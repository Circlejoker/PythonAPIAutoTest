"""
读取excel
用到了  1. openpyxl库里的 load_workbook方法    如果没有安装 直接 pip install openpyxl  即可
       2. zip函数  ——> 可以将1对1的数据,组合起来以字典格式输出

"""
from openpyxl import load_workbook

class MyExcel:
    def __init__(self,excel_path,sheet_name):
        # 通过load_workbook函数，由输入的路径打开一个workbook
        wb =load_workbook(excel_path)
        # 打开一个表单 - 通过sheet页名
        self.sh = wb[sheet_name]
        pass

    def read_data(self):
        # 注：接口的请求数据，读取出来是字符串
        # 存储表单下读取到的所有数据 - 每个成员患都是一个字典
        all_data=[]
        data = list(self.sh.values)
        keys =data[0]
        for row in data[1:]:
            row_dict = dict(zip(keys,row))
            all_data.append(row_dict)
        return all_data